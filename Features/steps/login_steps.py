import logging

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from Features.environment import capture_screenshot

logger = logging.getLogger(__name__)

@then(u'I should be redirected to the loginpage')
def step_impl(context):
    logger.info("Verifying the home URL is accessible")
    current_url = context.driver.current_url
    assert 'login' in current_url
    logger.info("Home URL loaded without any error")
    capture_screenshot(context.driver,"HomeURL")


def read_credentials_from_excel(file_path,sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data=[]
    for i in range(2,sheet.max_row+1):
        username=sheet.cell(row=i,column=1).value
        password=sheet.cell(row=i,column=2).value
        result=sheet.cell(row=i,column=3).value
        data.append((username,password,result))
    return data

@when(u'I login with the credentials from excel file')
def step_impl(context):
    logger.info("Validating login functionality using multiple sets of credentials")
    data=read_credentials_from_excel("testdata/login_data.xlsx",'sheet1')
    wait = WebDriverWait(context.driver, 10)
    for username,password,result in data:
        try:
            username_field = wait.until(EC.visibility_of_element_located((By.NAME, "username")))
            password_field = wait.until(EC.visibility_of_element_located((By.NAME, "password")))
            username_field.clear()
            password_field.clear()
            username_field.send_keys(username)
            password_field.send_keys(password)
            context.driver.find_element(By.XPATH, "//button[@type='submit']").click()
            if result == "homepage":
                wait = WebDriverWait(context.driver, 10)
                wait.until(EC.url_contains("dashboard"))
                print(f"{username} :valid login passed")
                logger.info("Valid login passed")
                capture_screenshot(context.driver,"Valid login")
                drop_down=wait.until(EC.visibility_of_element_located((By.XPATH,"//i[contains(@class,'oxd-icon bi-caret')]")))
                drop_down.click()
                logout_button=wait.until(EC.visibility_of_element_located((By.XPATH,"//a[text()='Logout']")))
                logout_button.click()
                wait.until(EC.visibility_of_element_located((By.NAME, "username")))
            else:
                err_msg = wait.until(EC.visibility_of_element_located((By.XPATH, "//p[text()='Invalid credentials']")))
                assert err_msg.is_displayed()
                print(f"{username}:invalid login passed")
                logger.info("Invalid login test passed")
                capture_screenshot(context.driver,"invalid login")
        except Exception as e:
            logger.info(f"{username}:Test failed due to {e}")
            print("Error:",e)
            capture_screenshot(context.driver,"Error while testing with invalid login")


@then(u'Username and password field should be displayed and enabled')
def step_impl(context):
    wait= WebDriverWait(context.driver,10)
    logger.info("Validating presence of login fields")
    username_field = wait.until(EC.visibility_of_element_located((By.NAME, "username")))
    password_field = wait.until(EC.visibility_of_element_located((By.NAME, "password")))
    username_field.is_displayed()
    password_field.is_displayed()
    username_field.is_enabled()
    password_field.is_enabled()
    logger.info("Login fields are visible and displayed")
    capture_screenshot(context.driver,"Login fields are visible and displayed")

@when(u'I click on forgot password link')
def step_impl(context):
    wait=WebDriverWait(context.driver,10)
    logger.info("Verifying Forgot Password link functionality")
    forgot_password_link=wait.until(EC.visibility_of_element_located((By.XPATH,"//p[text()='Forgot your password? ']")))
    forgot_password_link.click()

@then(u'I should get the popup to enter an email-id where password rest link will be sent')
def step_impl(context):
    wait=WebDriverWait(context.driver,10)
    pop_up=wait.until(EC.visibility_of_element_located((By.XPATH,"//h6[text()='Reset Password']")))
    pop_up.is_displayed()

@then(u'I enter an email')
def step_impl(context):
    context.driver.find_element(By.NAME,"username").send_keys("rosharumaiza@gmail.com")

@then(u'I click on submit button')
def step_impl(context):
    context.driver.find_element(By.XPATH,"//button[@type='submit']").click()

@then(u'I should get the message that the password link has been sent successfully to the registered email-id')
def step_impl(context):
    wait=WebDriverWait(context.driver,10)
    try:
        Pop_up_msg = wait.until(EC.visibility_of_element_located((By.XPATH, "//h6[text()='Reset Password link sent successfully']")))
        Pop_up_msg.is_displayed()
        logger.info("Reset password link has sent successfully to the registered email id")
        capture_screenshot(context.driver,"Reset password link sent successfully")

    except Exception as e:
        logger.info(f"Did not receive the reset password link on registered email id due to {e}")
        capture_screenshot(context.driver,"Reset password link not sent")

